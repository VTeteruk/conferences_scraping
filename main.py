import logging
import re
import PyPDF2
from openpyxl.reader.excel import load_workbook
from utilities import delete_numbers_from_text, clean_text_for_names

# Set up logging
logging.basicConfig(level=logging.INFO)


class BookScraper:
    def __init__(self, pdf_file_path: str, excel_file_path: str) -> None:
        # Define field names for the data
        self.data_field_names = (
            "name",
            "affiliation",
            "location",
            "session",
            "topic",
            "presentation",
        )
        self.pdf_file_path = pdf_file_path
        self.excel_file_path = excel_file_path

    def extract_text_from_pdf(self, start_page: int = 0, end_page: int = None) -> str:
        # Create a PdfFileReader object
        pdf_reader = PyPDF2.PdfReader(self.pdf_file_path)

        # Set the end page if not specified
        if end_page is None:
            end_page = len(pdf_reader.pages) - 4  # Last 4 pages are useless

        # Extract text from specified pages and join it into a single string
        return " ".join(
            [pdf_reader.pages[i].extract_text() for i in range(start_page, end_page)]
        )

    @staticmethod
    def clean_text(text: str) -> str:
        # Clean the text by replacing specific patterns with spaces or newline characters
        return (
            text.replace("\t", " ")
            .replace("\n-", "")
            .replace("5th World Psoriasis & Psoriatic Arthritis Conference 2018", "")
            .replace("Acta Derm Venereol 2018", "")
            .replace("Poster abstracts", "")
            .replace("www.medicaljournals.se/acta", "")
            .replace(".P", ".\nP")
            .replace("Background", "Introduction")
            .replace("\xad", " ")
            .replace("Introduction", "\nIntroduction")
            .replace("1,2", "\n")
        )

    @staticmethod
    def divide_text_by_session_name(text: str) -> list[str]:
        # Use regular expression to find session names (e.g., P146)
        session_names = re.findall("P[0-9]{3}", text)
        sections = []
        start_index = 0

        for session_name in session_names:
            end_index = text.find(session_name, start_index)

            if end_index != -1:
                section_text = text[start_index:end_index].strip()
                sections.append(section_text)
                start_index = end_index

        last_section = text[start_index:].strip()
        sections.append(last_section)

        # Remove sections with less than 3 characters (page numbers)
        return [section for section in sections if len(section) > 3]

    @staticmethod
    def find_topic(data: list) -> tuple[str, int, str | None]:
        topic = ""
        for index, text in enumerate(data):
            if text.isupper():
                topic += text
            else:
                # Check if it wasn't separated correctly
                only_upper_text = re.sub("[a-z]", "+", text).split("+")
                length_of_upper_text = len(only_upper_text[0])

                # Check if it is only uppercase text
                if len(only_upper_text[0].replace("-", " ").split()[0]) > 1:
                    topic += only_upper_text[0][:-1]
                    return topic, index + 1, text[length_of_upper_text - 1:]
                else:
                    return topic, index + 1, None

    @staticmethod
    def find_name(data: list) -> tuple[str, int]:
        name = []
        for index in range(0, len(data) - 1):
            text = clean_text_for_names(data[index])
            future_text = clean_text_for_names(data[index + 1])

            name.append(data[index])
            if not (text.endswith(" ") or future_text.startswith(",")):
                return "".join(name), index + 1

    @staticmethod
    def split_location_and_presentation(data: list) -> list[str, str]:
        if "Introduction" in "".join(data):
            return "".join(data).split("Introduction")
        else:
            return [data[0], "".join(data[1:])]

    @staticmethod
    def get_affiliation_from_full_location(full_location: str) -> tuple[str, str]:
        split_location = full_location.split(", ")
        if len(split_location) == 1:
            affiliation = split_location
            location = ""
        else:
            *affiliation, location = split_location

        return " ".join(affiliation), location

    def parse_section(self, section: str) -> dict:
        # Split the section into lines
        split_data = section.split("\n")
        session = split_data[0].strip()

        # Find the topic and starting index
        topic, start_index, possible_name = self.find_topic(split_data[1:])

        if possible_name:
            split_data[start_index] = possible_name

        name_and_index = self.find_name(split_data[start_index:])
        name = name_and_index[0]

        start_index += name_and_index[1]

        full_location, presentation = self.split_location_and_presentation(
            split_data[start_index:]
        )
        affiliation, location = self.get_affiliation_from_full_location(full_location)

        presentation = (
            "Introduction" if presentation.startswith(":") else "Introduction "
        ) + presentation

        logging.info(f"Session {session} was parsed")

        return dict(
            zip(
                self.data_field_names,
                (
                    delete_numbers_from_text(name),
                    delete_numbers_from_text(affiliation),
                    delete_numbers_from_text(location),
                    session,
                    presentation,
                ),
            )
        )

    def add_data_to_excel(self, articles: list, start_row: int):
        # Load the Excel workbook
        wb = load_workbook(self.excel_file_path)

        # Select the active sheet
        ws = wb.active

        # Iterate through articles and add data to the Excel sheet
        for i, article in enumerate(articles, start=start_row):
            for column_num, column_name in enumerate(self.data_field_names, start=1):
                ws.cell(row=i, column=column_num, value=article.get(column_name))

        # Save the Excel workbook
        wb.save(self.excel_file_path)

        logging.info("Data was added to the Excel file successfully")


if __name__ == "__main__":
    pdf_file_to_parse = "data_to_parse/Book.pdf"
    excel_file_to_fill = (
        "results/Data Entry - 5th World Psoriasis & "
        "Psoriatic Arthritis Conference 2018 - Case format (2).xlsx"
    )

    # Create an instance of the BookScraper class
    book_scraper = BookScraper(
        pdf_file_path=pdf_file_to_parse, excel_file_path=excel_file_to_fill
    )

    # Extract text from the PDF
    pdf_text = book_scraper.extract_text_from_pdf(start_page=43)

    # Clean the extracted text
    cleaned_text = book_scraper.clean_text(pdf_text)

    # Divide the text into sections based on session names
    all_parts = book_scraper.divide_text_by_session_name(cleaned_text)

    # Parse each section to extract information
    parsed_articles = [book_scraper.parse_section(part) for part in all_parts]

    # Add parsed data to an Excel file
    book_scraper.add_data_to_excel(articles=parsed_articles, start_row=7)
